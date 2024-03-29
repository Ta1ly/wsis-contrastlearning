
import pandas as pd
from openpyxl import load_workbook
import os

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None
    Usage examples:
    >>> append_df_to_excel('d:/temp/test.xlsx', df)
    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)
    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)
    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)
    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

# def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
#                        truncate_sheet=False, 
#                        **to_excel_kwargs):
#     """
#     Append a DataFrame [df] to existing Excel file [filename]
#     into [sheet_name] Sheet.
#     If [filename] doesn't exist, then this function will create it.

#     Parameters:
#       filename : File path or existing ExcelWriter
#                  (Example: '/path/to/file.xlsx')
#       df : dataframe to save to workbook
#       sheet_name : Name of sheet which will contain DataFrame.
#                    (default: 'Sheet1')
#       startrow : upper left cell row to dump data frame.
#                  Per default (startrow=None) calculate the last row
#                  in the existing DF and write to the next row...
#       truncate_sheet : truncate (remove and recreate) [sheet_name]
#                        before writing DataFrame to Excel file
#       to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
#                         [can be dictionary]

#     Returns: None

#     (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
#     codes: https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
#     """
    

#     # ignore [engine] parameter if it was passed
#     if 'engine' in to_excel_kwargs:
#         to_excel_kwargs.pop('engine')

#     writer = pd.ExcelWriter(filename, engine='openpyxl')

#     # # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
#     # try:
#     #     FileNotFoundError
#     # except NameError:
#     #     FileNotFoundError = IOError

#     if os.path.exists(filename):
#         print('exist')
#     else:
#         print('not exist')
#     try:
#         # try to open an existing workbook
#         writer.book = load_workbook(filename)
        
#         # get the last row in the existing Excel sheet
#         # if it was not specified explicitly
#         if startrow is None and sheet_name in writer.book.sheetnames:
#             startrow = writer.book[sheet_name].max_row

#         # truncate sheet
#         if truncate_sheet and sheet_name in writer.book.sheetnames:
#             # index of [sheet_name] sheet
#             idx = writer.book.sheetnames.index(sheet_name)
#             # remove [sheet_name]
#             writer.book.remove(writer.book.worksheets[idx])
#             # create an empty sheet [sheet_name] using old index
#             writer.book.create_sheet(sheet_name, idx)
        
#         # copy existing sheets
#         writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
#     except FileNotFoundError:
#         # file does not exist yet, we will create it
#         pass

#     if startrow is None:
#         startrow = 0

#     # write out the new sheet
#     df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

#     # save the workbook
#     writer.save()